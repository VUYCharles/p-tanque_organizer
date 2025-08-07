import random
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class Joueur:
    def __init__(self, id, nom, role):
        self.id = id
        self.nom = nom
        self.role = role  # 'tireur', 'pointeur', 'milieu'

    def __repr__(self):
        return f"{self.nom} ({self.role})"

class Equipe:
    def __init__(self, tireur, pointeur, milieu, id_equipe):
        self.tireur = tireur
        self.pointeur = pointeur
        self.milieu = milieu
        self.id = id_equipe

    def __repr__(self):
        return f"Équipe {self.id}: [T: {self.tireur.nom}, P: {self.pointeur.nom}, M: {self.milieu.nom}]"

class Match:
    def __init__(self, equipe1, equipe2):
        self.equipe1 = equipe1
        self.equipe2 = equipe2

    def __repr__(self):
        return f"{self.equipe1} vs {self.equipe2}"

class Tournoi:
    def __init__(self, joueurs):
        self.joueurs = joueurs
        self.rounds = []
        self.historique_equipes = []
        self.historique_matchs_tireurs = set()
        self.historique_compositions = set()

    def generer_equipes(self):
        tireurs = [j for j in self.joueurs if j.role == 'tireur']
        pointeurs = [j for j in self.joueurs if j.role == 'pointeur']
        milieux = [j for j in self.joueurs if j.role == 'milieu']

        random.shuffle(tireurs)
        random.shuffle(pointeurs)
        random.shuffle(milieux)

        equipes = []
        for i in range(8):
            equipe = Equipe(tireurs[i], pointeurs[i], milieux[i], i+1)
            equipes.append(equipe)
        return equipes

    def est_equipe_valide(self, equipes):
        composition_actuelle = set()
        for equipe in equipes:
            composition = (equipe.tireur.id, equipe.pointeur.id, equipe.milieu.id)
            if composition in self.historique_compositions:
                return False
            composition_actuelle.add(composition)
        self.historique_compositions.update(composition_actuelle)
        return True

    def generer_matchs(self, equipes):
        paires = list(combinations(equipes, 2))
        random.shuffle(paires)

        matchs = []
        equipes_utilisees = set()
        nouveaux_matchs_tireurs = set()

        for equipe1, equipe2 in paires:
            if equipe1.id in equipes_utilisees or equipe2.id in equipes_utilisees:
                continue

            paire_tireurs = frozenset({equipe1.tireur.id, equipe2.tireur.id})
            if paire_tireurs not in self.historique_matchs_tireurs:
                matchs.append(Match(equipe1, equipe2))
                equipes_utilisees.add(equipe1.id)
                equipes_utilisees.add(equipe2.id)
                nouveaux_matchs_tireurs.add(paire_tireurs)

        equipes_restantes = [e for e in equipes if e.id not in equipes_utilisees]
        while len(equipes_restantes) >= 2:
            equipe1 = equipes_restantes.pop()
            equipe2 = equipes_restantes.pop()
            paire_tireurs = frozenset({equipe1.tireur.id, equipe2.tireur.id})
            if paire_tireurs not in self.historique_matchs_tireurs:
                matchs.append(Match(equipe1, equipe2))
                nouveaux_matchs_tireurs.add(paire_tireurs)

        self.historique_matchs_tireurs.update(nouveaux_matchs_tireurs)
        return matchs if len(matchs) == 4 else None

    def jouer_round(self, round_num):
        max_tentatives = 1000
        tentatives = 0

        while tentatives < max_tentatives:
            tentatives += 1
            equipes = self.generer_equipes()

            if not self.est_equipe_valide(equipes):
                continue

            matchs = self.generer_matchs(equipes)

            if matchs:
                self.historique_equipes.append(equipes)
                self.rounds.append(matchs)
                return True

        return False

    def jouer_tournoi(self):
        for i in range(4):
            if not self.jouer_round(i+1):
                return False
        return True

    def afficher_tournoi(self):
        for i, (round_matches, equipes) in enumerate(zip(self.rounds, self.historique_equipes), 1):
            print(f"\n=== Round {i} ===")
            print("\nComposition des équipes:")
            for equipe in equipes:
                print(f"  {equipe}")
            print("\nMatchs:")
            for j, match in enumerate(round_matches, 1):
                print(f"  Match {j}: {match}")

def generer_joueurs():
    joueurs = []
    roles = ['tireur']*8 + ['pointeur']*8 + ['milieu']*8

    for i in range(24):
        nom = f"Joueur_{i+1:02d}"
        role = roles[i]
        joueurs.append(Joueur(i+1, nom, role))

    return joueurs

def tester_contraintes(tournoi):
    print("\n=== Vérification des contraintes ===")

    assert len(tournoi.rounds) == 4
    print("✓ 4 rounds organisés")

    for round_equipes in tournoi.historique_equipes:
        assert len(round_equipes) == 8
    print("✓ 8 équipes par round")

    toutes_compositions = set()
    for round_equipes in tournoi.historique_equipes:
        for equipe in round_equipes:
            composition = (equipe.tireur.id, equipe.pointeur.id, equipe.milieu.id)
            assert composition not in toutes_compositions
            toutes_compositions.add(composition)
    print("✓ Aucune équipe identique entre les rounds")

    confrontations_tireurs = set()
    for round_matches in tournoi.rounds:
        for match in round_matches:
            paire = frozenset({match.equipe1.tireur.id, match.equipe2.tireur.id})
            assert paire not in confrontations_tireurs
            confrontations_tireurs.add(paire)
    print("✓ Aucun tireur ne rencontre le même adversaire plus d'une fois")

    for round_equipes in tournoi.historique_equipes:
        joueurs_round = set()
        for equipe in round_equipes:
            joueurs_round.add(equipe.tireur.id)
            joueurs_round.add(equipe.pointeur.id)
            joueurs_round.add(equipe.milieu.id)
        assert len(joueurs_round) == 24
    print("✓ Tous les joueurs participent à chaque round")

    for round_equipes in tournoi.historique_equipes:
        for equipe in round_equipes:
            assert equipe.tireur.role == 'tireur'
            assert equipe.pointeur.role == 'pointeur'
            assert equipe.milieu.role == 'milieu'
    print("✓ Composition correcte des équipes")

    for round_matches in tournoi.rounds:
        assert len(round_matches) == 4
    print("✓ 4 matchs par round")

    print("\n✓✓✓ Toutes les contraintes sont respectées ! ✓✓✓")

def generer_excel(tournoi, filename="tournoi_petanque.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tournoi"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:F1')
    ws['A1'] = "TOURNOI DE PÉTANQUE - COMPOSITION DES ÉQUIPES ET MATCHS"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment

    for round_num, (equipes, matches) in enumerate(zip(tournoi.historique_equipes, tournoi.rounds), 1):
        start_row = 3 + (round_num-1)*30
        ws.merge_cells(f'A{start_row}:F{start_row}')
        ws[f'A{start_row}'] = f"ROUND {round_num}"
        ws[f'A{start_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{start_row}'].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        ws[f'A{start_row}'].alignment = center_alignment

        headers = ["Équipe", "Tireur", "Pointeur", "Milieu"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        for i, equipe in enumerate(equipes, 1):
            row = start_row + 1 + i
            ws.cell(row=row, column=1, value=f"Équipe {equipe.id}").border = border
            ws.cell(row=row, column=2, value=equipe.tireur.nom).border = border
            ws.cell(row=row, column=3, value=equipe.pointeur.nom).border = border
            ws.cell(row=row, column=4, value=equipe.milieu.nom).border = border

        matchs_row = start_row + 11
        ws.merge_cells(f'A{matchs_row}:F{matchs_row}')
        ws[f'A{matchs_row}'] = f"MATCHS DU ROUND {round_num}"
        ws[f'A{matchs_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{matchs_row}'].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        ws[f'A{matchs_row}'].alignment = center_alignment

        headers_matchs = ["Match", "Équipe 1", "vs", "Équipe 2"]
        for col, header in enumerate(headers_matchs, 1):
            cell = ws.cell(row=matchs_row+1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        for i, match in enumerate(matches, 1):
            row = matchs_row + 1 + i
            ws.cell(row=row, column=1, value=f"Match {i}").border = border
            ws.cell(row=row, column=2, value=f"Équipe {match.equipe1.id}\n({match.equipe1.tireur.nom})").border = border
            ws.cell(row=row, column=3, value="vs").border = border
            ws.cell(row=row, column=4, value=f"Équipe {match.equipe2.id}\n({match.equipe2.tireur.nom})").border = border
            ws.cell(row=row, column=2).alignment = center_alignment
            ws.cell(row=row, column=4).alignment = center_alignment

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25

    ws_confrontations = wb.create_sheet("Confrontations Tireurs")
    ws_confrontations.merge_cells('A1:C1')
    ws_confrontations['A1'] = "CONFRONTATIONS ENTRE TIREURS"
    ws_confrontations['A1'].font = Font(bold=True, size=14)
    ws_confrontations['A1'].alignment = center_alignment

    headers = ["Tireur 1", "Tireur 2", "Round"]
    for col, header in enumerate(headers, 1):
        cell = ws_confrontations.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    row = 4
    for round_num, matches in enumerate(tournoi.rounds, 1):
        for match in matches:
            ws_confrontations.cell(row=row, column=1, value=match.equipe1.tireur.nom).border = border
            ws_confrontations.cell(row=row, column=2, value=match.equipe2.tireur.nom).border = border
            ws_confrontations.cell(row=row, column=3, value=f"Round {round_num}").border = border
            row += 1

    for col in range(1, 4):
        ws_confrontations.column_dimensions[get_column_letter(col)].width = 25

    wb.save(filename)
    print(f"\nFichier Excel généré avec succès : {filename}")

def main():
    joueurs = generer_joueurs()
    print("Organisation du tournoi de pétanque...")
    tournoi = Tournoi(joueurs)

    if tournoi.jouer_tournoi():
        tournoi.afficher_tournoi()
        tester_contraintes(tournoi)
        generer_excel(tournoi)
    else:
        print("Impossible d'organiser le tournoi avec les contraintes données.")

if __name__ == "__main__":
    main()
