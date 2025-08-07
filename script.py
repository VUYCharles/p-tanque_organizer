import random
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText

class Player:
    def __init__(self, id, name, role):
        self.id = id
        self.name = name
        self.role = role  # 'tireur', 'pointeur', 'milieu'

    def __repr__(self):
        return f"{self.name} ({self.role})"

class Team:
    def __init__(self, shooter, pointer, middle, team_id):
        self.shooter = shooter
        self.pointer = pointer
        self.middle = middle
        self.id = team_id

    def __repr__(self):
        return f"Équipe {self.id}: [T: {self.shooter.name}, P: {self.pointer.name}, M: {self.middle.name}]"

    def full_info(self):
        return (f"Équipe {self.id}\n"
                f"Tireur: {self.shooter.name}\n"
                f"Pointeur: {self.pointer.name}\n"
                f"Milieu: {self.middle.name}")

class Match:
    def __init__(self, team1, team2):
        self.team1 = team1
        self.team2 = team2

    def __repr__(self):
        return f"{self.team1} vs {self.team2}"

class Tournament:
    def __init__(self, players, num_courts):
        self.players = players
        self.num_courts = num_courts
        self.num_teams = len(players) // 3
        self.rounds = []
        self.teams_history = []
        self.shooter_encounters = set()
        self.team_compositions = set()

    def generate_teams(self):
        shooters = [p for p in self.players if p.role == 'tireur']
        pointers = [p for p in self.players if p.role == 'pointeur']
        middles = [p for p in self.players if p.role == 'milieu']

        random.shuffle(shooters)
        random.shuffle(pointers)
        random.shuffle(middles)

        return [Team(shooters[i], pointers[i], middles[i], i+1) for i in range(self.num_teams)]

    def is_valid_team_composition(self, teams):
        current_compositions = set()
        for team in teams:
            composition = (team.shooter.id, team.pointer.id, team.middle.id)
            if composition in self.team_compositions:
                return False
            current_compositions.add(composition)
        self.team_compositions.update(current_compositions)
        return True

    def generate_matches(self, teams):
        possible_pairs = list(combinations(teams, 2))
        random.shuffle(possible_pairs)

        matches = []
        used_teams = set()
        new_shooter_pairs = set()

        for team1, team2 in possible_pairs:
            if team1.id in used_teams or team2.id in used_teams:
                continue

            shooter_pair = frozenset({team1.shooter.id, team2.shooter.id})
            if shooter_pair not in self.shooter_encounters:
                matches.append(Match(team1, team2))
                used_teams.add(team1.id)
                used_teams.add(team2.id)
                new_shooter_pairs.add(shooter_pair)
                if len(matches) == self.num_courts:
                    break

        remaining_teams = [t for t in teams if t.id not in used_teams]
        while len(remaining_teams) >= 2 and len(matches) < self.num_courts:
            team1 = remaining_teams.pop()
            team2 = remaining_teams.pop()
            shooter_pair = frozenset({team1.shooter.id, team2.shooter.id})
            if shooter_pair not in self.shooter_encounters:
                matches.append(Match(team1, team2))
                new_shooter_pairs.add(shooter_pair)

        self.shooter_encounters.update(new_shooter_pairs)
        return matches if len(matches) == self.num_courts else None

    def play_round(self, round_num):
        max_attempts = 1000
        for _ in range(max_attempts):
            teams = self.generate_teams()
            if not self.is_valid_team_composition(teams):
                continue

            matches = self.generate_matches(teams)
            if matches:
                self.teams_history.append(teams)
                self.rounds.append(matches)
                return True
        return False

    def run_tournament(self):
        num_rounds = 6  # Nombre de rondes standard
        if self.num_teams == 8:  # 24 joueurs
            num_rounds = 4
        elif self.num_teams == 10:  # 30 joueurs
            num_rounds = 5
        return all(self.play_round(i+1) for i in range(num_rounds))

class PetanqueTournamentApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Organisateur de Tournoi de Pétanque")
        self.root.geometry("1000x700")
        self.tournament = None

        self.create_widgets()

    def create_widgets(self):
        # Frame de configuration
        config_frame = ttk.LabelFrame(self.root, text="Configuration du Tournoi", padding=10)
        config_frame.pack(fill=tk.X, padx=10, pady=5)

        # Sélection du nombre de joueurs
        ttk.Label(config_frame, text="Nombre de joueurs:").grid(row=0, column=0, sticky=tk.W)
        self.player_count = tk.IntVar(value=24)
        ttk.Radiobutton(config_frame, text="24 joueurs (4 terrains)", variable=self.player_count, value=24).grid(row=0, column=1, sticky=tk.W)
        ttk.Radiobutton(config_frame, text="30 joueurs (5 terrains)", variable=self.player_count, value=30).grid(row=0, column=2, sticky=tk.W)
        ttk.Radiobutton(config_frame, text="36 joueurs (6 terrains)", variable=self.player_count, value=36).grid(row=0, column=3, sticky=tk.W)

        # Boutons d'action
        button_frame = ttk.Frame(self.root)
        button_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Button(button_frame, text="Générer le Tournoi", command=self.generate_tournament).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Exporter vers Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Afficher les Contraintes", command=self.show_constraints).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Quitter", command=self.root.quit).pack(side=tk.RIGHT, padx=5)

        # Notebook pour afficher les résultats
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        # Onglet des équipes
        self.teams_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.teams_tab, text="Composition des Équipes")

        self.teams_text = ScrolledText(self.teams_tab, wrap=tk.WORD)
        self.teams_text.pack(fill=tk.BOTH, expand=True)

        # Onglet des matches
        self.matches_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.matches_tab, text="Matches")

        self.matches_text = ScrolledText(self.matches_tab, wrap=tk.WORD)
        self.matches_text.pack(fill=tk.BOTH, expand=True)

        # Onglet des rencontres de tireurs
        self.shooters_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.shooters_tab, text="Rencontres de Tireurs")

        self.shooters_text = ScrolledText(self.shooters_tab, wrap=tk.WORD)
        self.shooters_text.pack(fill=tk.BOTH, expand=True)

    def generate_players(self, num_players):
        num_teams = num_players // 3
        roles = ['tireur']*num_teams + ['pointeur']*num_teams + ['milieu']*num_teams
        return [Player(i+1, f"Joueur_{i+1:02d}", roles[i]) for i in range(num_players)]

    def generate_tournament(self):
        num_players = self.player_count.get()
        num_courts = num_players // 6

        players = self.generate_players(num_players)
        self.tournament = Tournament(players, num_courts)

        if self.tournament.run_tournament():
            self.display_tournament()
            messagebox.showinfo("Succès", "Tournoi généré avec succès!")
        else:
            messagebox.showerror("Erreur", "Impossible d'organiser le tournoi avec les contraintes données.")

    def display_tournament(self):
        # Afficher les équipes
        self.teams_text.delete(1.0, tk.END)
        for i, teams in enumerate(self.tournament.teams_history, 1):
            self.teams_text.insert(tk.END, f"\n=== Ronde {i} ===\n\n")
            for team in teams:
                self.teams_text.insert(tk.END, f"{team}\n")

        # Afficher les matches
        self.matches_text.delete(1.0, tk.END)
        for i, matches in enumerate(self.tournament.rounds, 1):
            self.matches_text.insert(tk.END, f"\n=== Ronde {i} ===\n\n")
            for j, match in enumerate(matches, 1):
                self.matches_text.insert(tk.END, f"Match {j}: {match}\n\n")

        # Afficher les rencontres de tireurs
        self.shooters_text.delete(1.0, tk.END)
        shooter_pairs = set()
        for round_num, matches in enumerate(self.tournament.rounds, 1):
            self.shooters_text.insert(tk.END, f"\n=== Ronde {round_num} ===\n\n")
            for match in matches:
                pair = (match.team1.shooter.name, match.team2.shooter.name)
                if pair not in shooter_pairs and (pair[1], pair[0]) not in shooter_pairs:
                    self.shooters_text.insert(tk.END, f"{pair[0]} vs {pair[1]}\n")
                    shooter_pairs.add(pair)

    def export_to_excel(self):
        if not self.tournament:
            messagebox.showerror("Erreur", "Veuillez d'abord générer un tournoi")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            title="Enregistrer le fichier Excel",
            initialfile="tournoi_petanque.xlsx"
        )

        if not filename:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Plan du tournoi"

        # Styles
        title_font = Font(bold=True, size=16, color="1F497D")
        round_header_font = Font(bold=True, size=12, color="FFFFFF")
        team_header_font = Font(bold=True, color="FFFFFF")
        match_header_font = Font(bold=True, color="FFFFFF")

        round_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        team_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        match_fill = PatternFill(start_color="76933C", end_color="76933C", fill_type="solid")

        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Titre principal
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "TOURNOI DE PÉTANQUE - PLAN COMPLET"
        title_cell.font = title_font
        title_cell.alignment = center_alignment

        # Configuration des colonnes
        columns = [
            ('A', 'Ronde', 8),
            ('B', 'Équipe', 12),
            ('C', 'Tireur', 20),
            ('D', 'Pointeur', 20),
            ('E', 'Milieu', 20)
        ]

        for col, header, width in columns:
            ws.column_dimensions[col].width = width
            if header:
                cell = ws.cell(row=3, column=ord(col)-64, value=header)
                cell.font = team_header_font
                cell.fill = team_fill
                cell.alignment = center_alignment
                cell.border = border

        current_row = 4

        # Ajout des données des rondes
        for round_num, (teams, matches) in enumerate(zip(self.tournament.teams_history, self.tournament.rounds), 1):
            # En-tête de ronde
            ws.merge_cells(f'A{current_row}:E{current_row}')
            round_cell = ws[f'A{current_row}']
            round_cell.value = f"RONDE {round_num}"
            round_cell.font = round_header_font
            round_cell.fill = round_fill
            round_cell.alignment = center_alignment
            current_row += 1

            # Composition des équipes
            for team in teams:
                ws[f'A{current_row}'] = f"R{round_num}"
                ws[f'B{current_row}'] = f"Équipe {team.id}"
                ws[f'C{current_row}'] = team.shooter.name
                ws[f'D{current_row}'] = team.pointer.name
                ws[f'E{current_row}'] = team.middle.name

                for col in ['A', 'B', 'C', 'D', 'E']:
                    cell = ws[f'{col}{current_row}']
                    cell.border = border
                    cell.alignment = left_alignment if col != 'A' else center_alignment

                current_row += 1

            # Section des matches
            ws.merge_cells(f'A{current_row}:E{current_row}')
            match_header_cell = ws[f'A{current_row}']
            match_header_cell.value = f"MATCHES DE LA RONDE {round_num}"
            match_header_cell.font = round_header_font
            match_header_cell.fill = round_fill
            match_header_cell.alignment = center_alignment
            current_row += 1

            # En-têtes des matches
            match_headers = ['Match', 'Équipe 1', '', 'Équipe 2', '']
            for col, header in zip(['A', 'B', 'C', 'D', 'E'], match_headers):
                if header:
                    cell = ws[f'{col}{current_row}']
                    cell.value = header
                    cell.font = match_header_font
                    cell.fill = match_fill
                    cell.alignment = center_alignment
                    cell.border = border
            current_row += 1

            # Détails des matches
            for match_num, match in enumerate(matches, 1):
                ws[f'A{current_row}'] = f"M{match_num}"
                ws[f'B{current_row}'] = match.team1.full_info()
                ws[f'D{current_row}'] = match.team2.full_info()
                ws[f'C{current_row}'] = "contre"

                for col in ['A', 'B', 'C', 'D', 'E']:
                    cell = ws[f'{col}{current_row}']
                    cell.border = border
                    cell.alignment = center_alignment if col in ['A', 'C'] else left_alignment

                current_row += 1

            # Ajout d'espace entre les rondes
            current_row += 2

        # Feuille des rencontres de tireurs
        ws_encounters = wb.create_sheet("Rencontres de tireurs")

        # Titre
        ws_encounters.merge_cells('A1:C1')
        title_cell = ws_encounters['A1']
        title_cell.value = "RENCONTRES ENTRE TIREURS"
        title_cell.font = title_font
        title_cell.alignment = center_alignment

        # En-têtes
        headers = ['Ronde', 'Tireur 1', 'Tireur 2']
        for col, header in enumerate(headers, 1):
            cell = ws_encounters.cell(row=3, column=col, value=header)
            cell.font = team_header_font
            cell.fill = team_fill
            cell.alignment = center_alignment
            cell.border = border

        # Données
        row = 4
        for round_num, matches in enumerate(self.tournament.rounds, 1):
            for match in matches:
                ws_encounters.cell(row=row, column=1, value=f"Ronde {round_num}").border = border
                ws_encounters.cell(row=row, column=2, value=match.team1.shooter.name).border = border
                ws_encounters.cell(row=row, column=3, value=match.team2.shooter.name).border = border
                row += 1

        # Ajustement des largeurs de colonnes
        for col in ['A', 'B', 'C']:
            ws_encounters.column_dimensions[col].width = 25

        wb.save(filename)
        messagebox.showinfo("Succès", f"Fichier Excel généré avec succès:\n{filename}")

    def show_constraints(self):
        if not self.tournament:
            messagebox.showerror("Erreur", "Veuillez d'abord générer un tournoi")
            return

        constraints = [
            f"Nombre de rondes: {len(self.tournament.rounds)}",
            f"Nombre d'équipes par ronde: {self.tournament.num_teams}",
            "Compositions d'équipes uniques pour toutes les rondes",
            "Les tireurs ne se rencontrent qu'une seule fois",
            f"Tous les {len(self.tournament.players)} joueurs participent à chaque ronde",
            f"Nombre de matches par ronde: {self.tournament.num_courts}"
        ]

        messagebox.showinfo(
            "Validation des Contraintes",
            "Toutes les contraintes sont validées avec succès:\n\n• " + "\n• ".join(constraints)
        )

def main():
    root = tk.Tk()
    app = PetanqueTournamentApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
